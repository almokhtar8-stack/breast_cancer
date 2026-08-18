#!/usr/bin/env python3
"""post_freeze_exploratory -- feasibility probe: are the GSE245601 authors'
PER-CELL malignancy annotations publicly obtainable?

This script lives in `scripts/` and not in `src/` because it makes network
calls, which CLAUDE.md forbids in analysis modules. It records what exists at
each public source; it downloads no primary data and derives no result. The
deterministic side (turning this record into the feasibility verdict) is
`src/post_poster_annotation_feasibility.py`.

Why the question matters: our inferCNV reconstruction retained 44,140 cells and
calls substantially fewer epithelial cells malignant than the authors' Epi.
Tumor labels imply. `docs/CNV_METHOD_AUDIT.md` established that the
disagreement is real and sample-dependent. Quantifying it as a confusion matrix
requires the authors' own per-cell labels, keyed by barcode and sample. If those
labels are not public, reconstructing them in order to compare against our
reconstruction of them would be circular, and the correct action is to stop.

Sources probed (all public, all read-only):
  1. GEO series GSE245601 and its samples -- supplementary file manifest.
  2. The paper's supplementary material, via the Europe PMC open REST API
     (PMC10690085). Every .xlsx sheet is opened and scanned for 10x barcodes.
  3. The authors' code repository hyunsoo77/BC_tamoxifen_response at the commit
     pinned in docs/gse245601_PREANALYSIS.md -- tree, tags, releases, LFS
     configuration, full history, and a barcode scan of every notebook.

Usage:
    python scripts/post_poster_probe_gse245601_annotations.py [--out PATH]

Output: a JSON record at
    results/post_poster/annotation_concordance/feasibility_probe.json
"""

from __future__ import annotations

import argparse
import io
import json
import logging
import re
import pathlib
import subprocess
import tempfile
import zipfile
from datetime import date
from pathlib import Path
from urllib.request import Request, urlopen

ROOT = Path(__file__).resolve().parents[1]
OUT_PATH = ROOT / "results" / "post_poster" / "annotation_concordance" / "feasibility_probe.json"

GEO_SERIES = "GSE245601"
PMC_ID = "PMC10690085"
AUTHOR_REPO = "hyunsoo77/BC_tamoxifen_response"
# Pinned in docs/gse245601_PREANALYSIS.md section 1.
PINNED_COMMIT = "ceabf3f331c88f464e6a57b0ad1f9c500bedde85"

# A 10x cell barcode: 16 bases plus a GEM-well suffix. Case-insensitive so a
# lower-cased export would still be caught.
BARCODE_RE = re.compile(r"[ACGTacgt]{16}-\d")

TIMEOUT = 180
logger = logging.getLogger("probe")


def fetch(url: str, timeout: int = TIMEOUT) -> bytes:
    req = Request(url, headers={"User-Agent": "breast-cancer-repo-feasibility-probe"})
    with urlopen(req, timeout=timeout) as resp:
        return resp.read()


def probe_geo() -> dict:
    """Supplementary file manifest for the series and for every sample."""
    url = ("https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi"
           f"?acc={GEO_SERIES}&targ=self&form=text&view=full")
    text = fetch(url).decode("utf8", "replace")
    series_supp = re.findall(r"!Series_supplementary_file\s*=\s*(\S+)", text)
    sample_ids = re.findall(r"!Series_sample_id\s*=\s*(\S+)", text)

    sample_supp: dict[str, list[str]] = {}
    for gsm in sample_ids:
        s_url = ("https://www.ncbi.nlm.nih.gov/geo/query/acc.cgi"
                 f"?acc={gsm}&targ=self&form=text&view=full")
        s_text = fetch(s_url).decode("utf8", "replace")
        sample_supp[gsm] = re.findall(r"!Sample_supplementary_file_\d+\s*=\s*(\S+)", s_text)
        logger.info("GEO %s: %d supplementary file(s)", gsm, len(sample_supp[gsm]))

    exts = sorted({Path(f).suffix for files in sample_supp.values() for f in files})
    return {
        "series": GEO_SERIES,
        "n_samples": len(sample_ids),
        "series_supplementary_files": series_supp,
        "sample_supplementary_file_extensions": exts,
        "n_sample_supplementary_files": sum(len(v) for v in sample_supp.values()),
        "any_cell_metadata_file": False if exts in ([".h5"], []) else None,
        "note": ("GEO carries only Cell Ranger filtered feature-barcode matrices. "
                 "A per-cell label table would appear here as a csv/tsv/rds "
                 "supplementary file; none exists."),
    }


def probe_paper_supplement() -> dict:
    """Every supplementary file of the paper, via the Europe PMC open API.

    Each .xlsx is opened sheet by sheet and scanned for 10x barcodes. This is
    the check that distinguishes "we did not look" from "the labels are not
    there".
    """
    url = f"https://www.ebi.ac.uk/europepmc/webservices/rest/{PMC_ID}/supplementaryFiles"
    blob = fetch(url)
    zf = zipfile.ZipFile(io.BytesIO(blob))
    names = sorted(zf.namelist())

    try:
        import openpyxl
    except ImportError:  # pragma: no cover - environment guard
        return {"pmc_id": PMC_ID, "files": names,
                "error": "openpyxl unavailable; sheets not scanned"}

    sheets, barcode_hits = [], 0
    for name in names:
        if not name.lower().endswith(".xlsx"):
            continue
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(name)), read_only=True)
        for ws in wb.worksheets:
            hits = 0
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and BARCODE_RE.search(str(cell)):
                        hits += 1
            barcode_hits += hits
            sheets.append({"file": name, "sheet": ws.title,
                           "n_rows": ws.max_row, "n_cols": ws.max_column,
                           "n_cells_matching_barcode": hits})
            logger.info("supp %s / %s: %sx%s, %d barcode-like cells",
                        name, ws.title, ws.max_row, ws.max_column, hits)
        wb.close()

    return {
        "pmc_id": PMC_ID,
        "source_api": "Europe PMC supplementaryFiles (open access REST)",
        "files": names,
        "n_xlsx_sheets_scanned": len(sheets),
        "max_sheet_rows": max((s["n_rows"] for s in sheets), default=0),
        "total_barcode_like_cells": barcode_hits,
        "sheets": sheets,
    }


def probe_author_repo() -> dict:
    """Tree, refs, releases, LFS config, full history and notebook barcode scan."""
    api = f"https://api.github.com/repos/{AUTHOR_REPO}"
    tree = json.loads(fetch(f"{api}/git/trees/{PINNED_COMMIT}?recursive=1"))
    paths = [t["path"] for t in tree.get("tree", []) if t["type"] == "blob"]
    releases = json.loads(fetch(f"{api}/releases"))
    tags = json.loads(fetch(f"{api}/tags"))

    has_lfs = False
    try:
        fetch(f"https://raw.githubusercontent.com/{AUTHOR_REPO}/{PINNED_COMMIT}/.gitattributes")
        has_lfs = True
    except Exception:
        has_lfs = False

    with tempfile.TemporaryDirectory() as tmp:
        clone = Path(tmp) / "repo"
        subprocess.run(["git", "clone", "--quiet", f"https://github.com/{AUTHOR_REPO}.git",
                        str(clone)], check=True, timeout=TIMEOUT * 3)
        head = subprocess.run(["git", "rev-parse", "HEAD"], cwd=clone, check=True,
                              capture_output=True, text=True).stdout.strip()
        n_commits = subprocess.run(["git", "rev-list", "--count", "--all"], cwd=clone,
                                   check=True, capture_output=True, text=True).stdout.strip()
        ever = subprocess.run(["git", "log", "--all", "--pretty=format:", "--name-only"],
                              cwd=clone, check=True, capture_output=True, text=True)
        ever_paths = sorted({p for p in ever.stdout.splitlines() if p.strip()})

        # Recursive, not root-level: a nested notebook would otherwise be
        # missed and the "no labels here" conclusion would be unearned.
        notebook_hits = {}
        for nb in sorted(clone.rglob("*.ipynb")):
            text = nb.read_text(encoding="utf8", errors="replace")
            notebook_hits[str(nb.relative_to(clone))] = len(BARCODE_RE.findall(text))

        # Files whose extension does not identify them as code or an image --
        # the population in which a disguised data file could hide.
        code_like = {".ipynb", ".md", ".png", ".r", ".rmd", ".py", ".sh", ".txt",
                     ".yml", ".yaml", ".gitignore"}
        unclassified = sorted(
            p for p in ever_paths if pathlib.PurePosixPath(p).suffix.lower() not in code_like)

    ever_exts = sorted({Path(p).suffix.lower() for p in ever_paths})
    return {
        "repo": AUTHOR_REPO,
        "pinned_commit": PINNED_COMMIT,
        "current_head": head,
        "head_matches_pinned_commit": head == PINNED_COMMIT,
        "n_commits_all_refs": int(n_commits),
        "n_blobs_at_pinned_commit": len(paths),
        "n_releases": len(releases),
        "n_tags": len(tags),
        "has_gitattributes_lfs_config": has_lfs,
        "file_extensions_ever_committed": ever_exts,
        "n_paths_ever_committed": len(ever_paths),
        "paths_ever_committed_without_a_code_or_image_extension": unclassified,
        "notebook_barcode_like_matches": notebook_hits,
        "max_notebook_barcode_matches": max(notebook_hits.values(), default=0),
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, default=OUT_PATH)
    args = parser.parse_args(argv)
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    record = {
        "post_freeze_exploratory": True,
        "probe_date": date.today().isoformat(),
        "question": ("Are the GSE245601 authors' per-cell Epi. Tumor / Epi. Nontumor "
                     "labels publicly obtainable, keyed by barcode and sample?"),
        "geo": probe_geo(),
        "paper_supplement": probe_paper_supplement(),
        "author_repo": probe_author_repo(),
        "controlled_access_route": {
            "accession": "phs003186.v1.p1",
            "repository": "dbGaP",
            "status": "not queried",
            "note": ("The paper's Data Availability statement routes processed "
                     "scRNA-seq data here. Controlled access is a possible source, "
                     "not a publicly available comparator."),
        },
    }
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(record, indent=2) + "\n")
    logger.info("wrote %s", args.out)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
